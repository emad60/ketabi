# warehouses/excel_reports.py
"""
توليد تقارير Excel للإحصائيات
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from io import BytesIO


class ExcelReportGenerator:
    """مولد تقارير Excel"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Styles
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.title_font = Font(name='Arial', size=16, bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _set_column_width(self, column, width):
        """تعيين عرض العمود"""
        self.ws.column_dimensions[get_column_letter(column)].width = width
    
    def _add_title(self, title, row=1):
        """إضافة عنوان التقرير"""
        self.ws.merge_cells(f'A{row}:F{row}')
        cell = self.ws[f'A{row}']
        cell.value = title
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 30
    
    def _add_date(self, row=2):
        """إضافة تاريخ التقرير"""
        self.ws.merge_cells(f'A{row}:F{row}')
        cell = self.ws[f'A{row}']
        cell.value = f"تاريخ التقرير: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, italic=True)
    
    def _add_header_row(self, headers, row):
        """إضافة صف العناوين"""
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def _add_data_row(self, data, row):
        """إضافة صف بيانات"""
        for col, value in enumerate(data, 1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = value
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def generate_ministry_statistics_report(self, stats_data):
        """توليد تقرير إحصائيات الوزارة"""
        self._add_title('تقرير إحصائيات الوزارة الشامل')
        self._add_date()
        
        current_row = 4
        
        # إحصائيات عامة
        self.ws[f'A{current_row}'].value = 'الإحصائيات العامة'
        self.ws[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        general_headers = ['البند', 'العدد']
        self._add_header_row(general_headers, current_row)
        current_row += 1
        
        general_stats = [
            ['إجمالي المحافظات', stats_data.get('total_provinces', 0)],
            ['إجمالي المدارس', stats_data.get('total_schools', 0)],
            ['إجمالي الكتب', stats_data.get('total_books', 0)],
            ['إجمالي الشحنات', stats_data.get('total_shipments', 0)],
            ['المستودعات', stats_data.get('total_warehouses', 0)],
            ['المناديب', stats_data.get('total_drivers', 0)],
        ]
        
        for stat in general_stats:
            self._add_data_row(stat, current_row)
            current_row += 1
        
        current_row += 2
        
        # حالات الشحنات
        self.ws[f'A{current_row}'].value = 'حالات الشحنات'
        self.ws[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 1
        
        shipment_headers = ['الحالة', 'العدد', 'النسبة']
        self._add_header_row(shipment_headers, current_row)
        current_row += 1
        
        shipments_by_status = stats_data.get('shipments_by_status', {})
        total_shipments = sum(shipments_by_status.values()) or 1
        
        status_names = {
            'pending': 'قيد الإنشاء',
            'assigned': 'مُسندة',
            'out_for_delivery': 'خارجة للتسليم',
            'delivered': 'تم التسليم',
            'confirmed': 'مؤكدة',
            'canceled': 'ملغاة'
        }
        
        for status, count in shipments_by_status.items():
            percentage = (count / total_shipments) * 100
            self._add_data_row([
                status_names.get(status, status),
                count,
                f"{percentage:.1f}%"
            ], current_row)
            current_row += 1
        
        # ضبط عرض الأعمدة
        self._set_column_width(1, 30)
        self._set_column_width(2, 15)
        self._set_column_width(3, 15)
        
        return self.wb
    
    def generate_province_statistics_report(self, province_name, stats_data):
        """توليد تقرير إحصائيات محافظة"""
        self._add_title(f'تقرير إحصائيات محافظة {province_name}')
        self._add_date()
        
        current_row = 4
        
        # إحصائيات المحافظة
        headers = ['البند', 'العدد']
        self._add_header_row(headers, current_row)
        current_row += 1
        
        province_stats = [
            ['المستودعات', stats_data.get('warehouses_count', 0)],
            ['المدارس', stats_data.get('schools_count', 0)],
            ['المناديب', stats_data.get('drivers_count', 0)],
            ['الشحنات الواردة', stats_data.get('incoming_shipments', 0)],
            ['الشحنات الموزعة', stats_data.get('distributed_shipments', 0)],
            ['المخزون الحالي', stats_data.get('current_stock', 0)],
        ]
        
        for stat in province_stats:
            self._add_data_row(stat, current_row)
            current_row += 1
        
        self._set_column_width(1, 30)
        self._set_column_width(2, 15)
        
        return self.wb
    
    def generate_warehouse_stock_report(self, warehouse_name, stock_data):
        """توليد تقرير مخزون مستودع"""
        self._add_title(f'تقرير مخزون {warehouse_name}')
        self._add_date()
        
        current_row = 4
        
        headers = ['المادة', 'الصف', 'الفصل', 'الكمية', 'الحد الأدنى', 'الحالة']
        self._add_header_row(headers, current_row)
        current_row += 1
        
        for item in stock_data:
            status = 'منخفض' if item.get('is_low_stock') else 'جيد'
            self._add_data_row([
                item.get('book_subject', '-'),
                item.get('book_grade', '-'),
                item.get('term', '-'),
                item.get('quantity', 0),
                item.get('min_threshold', 0),
                status
            ], current_row)
            current_row += 1
        
        for col in range(1, 7):
            self._set_column_width(col, 20)
        
        return self.wb
    
    def generate_shipments_report(self, shipments_data):
        """توليد تقرير الشحنات"""
        self._add_title('تقرير الشحنات')
        self._add_date()
        
        current_row = 4
        
        headers = ['رقم الشحنة', 'من', 'إلى', 'المندوب', 'الحالة', 'التاريخ']
        self._add_header_row(headers, current_row)
        current_row += 1
        
        for shipment in shipments_data:
            self._add_data_row([
                shipment.get('tracking_code', '-'),
                shipment.get('from_name', '-'),
                shipment.get('to_name', '-'),
                shipment.get('courier_name', 'غير مُسند'),
                shipment.get('status_display', '-'),
                shipment.get('created_at', '-')
            ], current_row)
            current_row += 1
        
        for col in range(1, 7):
            self._set_column_width(col, 20)
        
        return self.wb
    
    def save_to_bytes(self):
        """حفظ الـ Workbook إلى BytesIO"""
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
